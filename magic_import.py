#!/usr/bin/env python
# -*- coding: utf-8 -*-

# This script relies on openpyxl. I don't want to install that as requirement for PLAID since this script will be run
# a few times tops. So remember to install it manually.

import argparse
import openpyxl
import os
import psycopg2
import subprocess
import sys

DEV_CONNECTION_STRING = "host='localhost' dbname='pid_db'"
PROD_CONNECTION_STRING = "host='prd-db' port='5433' dbname='plaid_db' user='plaid_user' password='your_password_here'"
STAGING_CONNECTION_STRING = "host='prd-db' port='5433' dbname='plaid_staging_db' user='plaid_staging_user' password='your_password_here'"


def main():
    parser = argparse.ArgumentParser(description='Magic Import for PID')
    parser.add_argument('data_file', help='Excel spreadsheet with all import data')
    parser.add_argument('-e', '--environment', choices=['dev', 'prod', 'staging'], default='dev', help='Specify which env db to use')
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.data_file, read_only=True)  # Open Excel file

    # Sanity check file, make sure all expected sheets and columns are there
    expected_sheets = populate_expected_sheets()
    data_sheets = populate_data_sheets(wb)
    if list(expected_sheets.keys()) != list(data_sheets.keys()):
        print('Excel sheets differ from expected sheets. Aborting.')
        print('Expected: {0}'.format(list(expected_sheets.keys())))
        print('Got:      {0}'.format(list(data_sheets.keys())))
        sys.exit(1)
    for key, values in data_sheets.items():
        if expected_sheets[key] != values:
            print('Excel headers for {0} differ from expected headers. Aborting.'.format(key))
            print('Expected: {0}'.format(expected_sheets[key]))
            print('Got:      {0}'.format(values))
            sys.exit(1)

    # Connect to DB
    if args.environment == 'prod':
        conn = psycopg2.connect(PROD_CONNECTION_STRING)
    elif args.environment == 'staging':
        conn = psycopg2.connect(STAGING_CONNECTION_STRING)
    else:
        conn = psycopg2.connect(DEV_CONNECTION_STRING)

    c = conn.cursor()  # get a cursor object used to execute SQL commands

    # Wipe current DB: Drop schema, delete migrations folder, manage.py db init, manage.py db migrate, manage.py.db upgrade
    c.execute('DROP SCHEMA public CASCADE')
    if args.environment == 'prod':
        if os.path.isdir('migrations_prod'):
            subprocess.call(['rm', '-rf', 'migrations_prod'])
    elif args.environment == 'staging':
        if os.path.isdir('migrations_staging'):
            subprocess.call(['rm', '-rf', 'migrations_staging'])
    else:
        if os.path.isdir('migrations_dev'):
            subprocess.call(['rm', '-rf', 'migrations_dev'])
    c.execute('CREATE SCHEMA public')
    if args.environment == 'prod':
        os.environ['PLAID_ENV'] = 'prod'
    if args.environment == 'staging':
        os.environ['PLAID_ENV'] = 'staging'
    conn.commit()  # Need to commit and close DB so alembic can create schema and tables
    conn.close()
    subprocess.call(['python', 'manage.py', 'db', 'init'])
    subprocess.call(['python', 'manage.py', 'db', 'migrate'])
    subprocess.call(['python', 'manage.py', 'db', 'upgrade'])

    # Connect to DB
    if args.environment == 'prod':
        conn = psycopg2.connect(PROD_CONNECTION_STRING)
    elif args.environment == 'staging':
        conn = psycopg2.connect(STAGING_CONNECTION_STRING)
    else:
        conn = psycopg2.connect(DEV_CONNECTION_STRING)
    c = conn.cursor()  # get a cursor object used to execute SQL commands

    # Sanity check DB, make sure all expected tables and columns are there
    expected_tables = populate_expected_tablestructure()
    data_tables = populate_data_tables(c, list(expected_tables.keys()))
    for key, values in data_tables.items():
        if expected_tables[key].sort() != values.sort():  # Have to .sort() as tables might shift around due to init new DB
            print('DB columns for {0} differ from expected columns. Aborting.'.format(key))
            print('Expected: {0}'.format(expected_tables[key]))
            print('Got:      {0}'.format(values))
            sys.exit(1)

    # Create sequences in use in app which don't get included by alembic
    c.execute('CREATE SEQUENCE anomaly_key_seq INCREMENT BY 1 MINVALUE 1 OWNED BY anomalies.key')  # TODO: no longer need for this if starts on 1, can just use primary key
    c.execute('CREATE SEQUENCE eco_key_seq INCREMENT BY 1 MINVALUE 1 OWNED BY ecos.key')  # TODO: no longer need for this if starts on 1, can just use primary key
    c.execute('CREATE SEQUENCE specification_number_seq INCREMENT BY 1 MINVALUE 10000 OWNED BY specifications.specification_number')
    c.execute('CREATE SEQUENCE procedure_number_seq INCREMENT BY 1 MINVALUE 300 OWNED BY procedures.procedure_number')
    c.execute('CREATE SEQUENCE task_number_seq INCREMENT BY 1 MINVALUE 1 OWNED BY tasks.task_number')  # TODO: no longer need for this if starts on 1, can just use primary key

    # Create a user for Sean, who will be owner for everything except specifications in magic import. Will always be ID 1
    c.execute("INSERT INTO users (first_name, last_name, username, email, roles) VALUES ('Sean', 'Haggart', 'sean', 'haggartplaid@gmail.com', 'plaid-users, plaid-superusers')")
    # Create a user for Peter, who will be owner for all specifications in magic import. Will always be ID 2
    c.execute("INSERT INTO users (first_name, last_name, username, email, roles) VALUES ('P', 'I', 'specs', 'specs@planetaryresources.com', 'plaid-users')")
    # Create a user for Jarle, mostly helpful for local testing where AD is not always reachable, ID 3
    c.execute("INSERT INTO users (first_name, last_name, username, email, roles) VALUES ('Jarle', 'Hakas', 'jarle', 'jarle@planetaryresources.com', 'plaid-users, plaid-admins')")
    # Create a user for Selena for EFAB ownership, ID 4
    c.execute("INSERT INTO users (first_name, last_name, username, email, roles) VALUES ('Selena', 'Oliver', 'selena', 'ssullivan@planetaryresources.com', 'plaid-users')")
    # Create a user for Charlie for MFAB ownership, ID 5
    c.execute("INSERT INTO users (first_name, last_name, username, email, roles) VALUES ('Charlie', 'Gary', 'cgary', 'charlie@planetaryresources.com', 'plaid-users')")

    # Start inserting data from Excel file
    # Do tables manually because some require some logic
    insert_common_tables(c, wb, 'CRITICALITY', 'criticalities')
    insert_common_tables(c, wb, 'DISPOSITION', 'dispositions')
    insert_common_tables(c, wb, 'HW TYPE', 'hardware_types')
    insert_projects(c, wb)
    insert_companies(c, wb)
    insert_materials(c, wb)
    insert_specifications(c, wb)
    insert_vendor_parts_products(c, wb)
    insert_designs(c, wb)

    # Create settings, Selena owner of EFAB, Charlie of MFAB and Jarle of PLAID
    c.execute("INSERT INTO plaid_settings (efab_user_id, mfab_user_id, plaid_admin_id, name_order) VALUES (4, 5, 3, 'last_name_first_name')")

    # Clear PLAID_ENV
    if os.environ.get('PLAID_ENV') is not None:
        del(os.environ['PLAID_ENV'])

    # Commit and close DB connection
    conn.commit()
    conn.close()


def insert_common_tables(c, wb, sheet_name, table):
    sheet = wb.get_sheet_by_name(sheet_name)
    for row in sheet.iter_rows(row_offset=1):
        name = row[0].value
        description = row[1].value
        ordering = row[2].value
        sql = 'INSERT INTO {0} (name, description, ordering) VALUES (%s, %s, %s)'.format(table)
        c.execute(sql, (name, description, ordering))


def insert_projects(c, wb):
    sheet = wb.get_sheet_by_name('PROJECT')
    for row in sheet.iter_rows(row_offset=1):
        name = row[0].value
        description = row[1].value
        sql = 'INSERT INTO projects (name, description) VALUES (%s, %s)'
        c.execute(sql, (name, description))


def insert_companies(c, wb):
    sheet = wb.get_sheet_by_name('COMPANY')
    for row in sheet.iter_rows(row_offset=1):
        name = row[0].value
        website = row[1].value
        pri_account_number = row[2].value
        terms = row[3].value
        alias = row[4].value
        address = row[5].value
        notes = row[6].value
        if name:
            sql = 'INSERT INTO companies (name, website, address, notes, alias, pri_account_number, terms) VALUES (%s, %s, %s, %s, %s, %s, %s)'
            c.execute(sql, (name, website, address, notes, alias, pri_account_number, terms))


def insert_materials(c, wb):
    sheet = wb.get_sheet_by_name('MATERIAL')
    for row in sheet.iter_rows(row_offset=1):
        name = row[0].value
        description = row[1].value
        ordering = row[6].value
        spec1 = row[2].value
        spec2 = row[3].value
        spec3 = row[4].value
        spec4 = row[5].value
        sql = 'INSERT INTO materials (name, description, ordering) VALUES (%s, %s, %s) RETURNING id'
        c.execute(sql, (name, description, ordering))
        material_id = c.fetchone()[0]
        if spec1:
            sql = 'INSERT INTO material_specifications (name, material_id) VALUES (%s, %s)'
            c.execute(sql, (spec1, material_id))
        if spec2:
            sql = 'INSERT INTO material_specifications (name, material_id) VALUES (%s, %s)'
            c.execute(sql, (spec2, material_id))
        if spec3:
            sql = 'INSERT INTO material_specifications (name, material_id) VALUES (%s, %s)'
            c.execute(sql, (spec3, material_id))
        if spec4:
            sql = 'INSERT INTO material_specifications (name, material_id) VALUES (%s, %s)'
            c.execute(sql, (spec4, material_id))


def insert_specifications(c, wb):
    sheet = wb.get_sheet_by_name('SPECS')
    revision = 'A'
    state = 'In Work'
    owner_id = 2  # No need for owner_name or anything, it's always Pete with ID 2
    created_by_id = 2  # Same as above
    for row in sheet.iter_rows(row_offset=1):
        # Create change log
        sql = 'INSERT INTO change_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        change_log_id = c.fetchone()[0]
        # Create first change log entry
        sql = 'INSERT INTO change_log_entries (parent_id, action, changed_by_id, changed_at) VALUES (%s, %s, %s, now()) RETURNING id'
        c.execute(sql, (change_log_id, 'Create', created_by_id))

        # Create workflow log
        sql = 'INSERT INTO workflow_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        workflow_log_id = c.fetchone()[0]
        # Create first workflow log entry
        sql = 'INSERT INTO workflow_log_entries (parent_id, action, capacity, changed_by_id, changed_at) VALUES (%s, %s, %s, %s, now()) RETURNING id'
        c.execute(sql, (workflow_log_id, 'Create', 'Owner', created_by_id))

        # Create revision log
        sql = 'INSERT INTO revision_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        revision_id = c.fetchone()[0]
        # Create first revision log entry
        sql = 'INSERT INTO revision_log_entries (parent_id, revision, reason, revisioned_by_id, revisioned_at) VALUES (%s, %s, %s, %s, now()) RETURNING id'
        c.execute(sql, (revision_id, 'A', 'Initial Release', 1))

        # Create specification
        spec_number = row[0].value
        title = row[1].value
        sql = 'INSERT INTO specifications (specification_number, revision, name, state, owner_id, created_by_id, created_at, change_log_id, workflow_log_id, revision_log_id) VALUES (%s, %s, %s, %s, %s, %s, now(), %s, %s, %s)'
        c.execute(sql, (spec_number, revision, title, state, owner_id, created_by_id, change_log_id, workflow_log_id, revision_id))


def insert_vendor_parts_products(c, wb):
    sheet = wb.get_sheet_by_name('VENDOR Parts.Products')
    # ['Part Number', 'Name / Description', 'Summary', 'Vendor', 'Owner', 'Project', 'CBE', 'UNC', 'Material', 'Matrl Spec', 'Build', 'TYPE', 'S/L #', 'Hardware Type', 'Project', 'Owner']
    # ['id', 'part_number', 'name', 'state', 'summary', 'notes', 'owner_id', 'created_by_id', 'created_at', 'project_id', 'vendor_id', 'thumbnail_id', 'current_best_estimate', 'material_id', 'material_specification_id', 'predicted_best_estimate', 'uncertainty']
    # ['id', 'serial_number', 'vendor_part_id', 'summary', 'notes', 'state', 'product_type', 'measured_mass', 'hardware_type_id', 'owner_id', 'created_by_id', 'created_at', 'project_id', 'vendor_build_id', 'thumbnail_id', 'change_log_id']
    for row in sheet.iter_rows(row_offset=1):
        part_number = row[0].value
        name = row[1].value
        summary = row[2].value
        vendor_name = row[3].value
        project_name = row[5].value
        cbe = row[6].value
        unc = row[7].value
        material_name = row[8].value
        material_spec_name = row[9].value
        build_identifier = row[10].value
        product_type = row[11].value
        serial_number = row[12].value
        hardware_type_name = row[13].value
        product_project_name = row[14].value

        # Get vendor ID
        sql = 'SELECT id FROM companies WHERE name = %s'
        c.execute(sql, (vendor_name,))
        vendor_id = c.fetchone()[0]
        # Get project ID
        sql = 'SELECT id FROM projects WHERE name = %s'
        c.execute(sql, (project_name,))
        project_id = c.fetchone()[0]
        # Get product_project ID
        sql = 'SELECT id FROM projects WHERE name = %s'
        c.execute(sql, (product_project_name,))
        product_project_id = c.fetchone()[0]
        # Get material ID
        sql = 'SELECT id FROM materials WHERE name = %s'
        c.execute(sql, (material_name,))
        material_id = c.fetchone()[0]
        # Get material_spec ID
        material_spec_id = None
        if material_spec_name:
            sql = 'SELECT id FROM material_specifications WHERE name = %s'
            c.execute(sql, (material_spec_name,))
            material_spec_id = c.fetchone()[0]
        # Get hardware type ID
        sql = 'SELECT id FROM hardware_types WHERE name = %s'
        c.execute(sql, (hardware_type_name,))
        hardware_type_id = c.fetchone()[0]

        # Create change log
        sql = 'INSERT INTO change_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        change_log_id = c.fetchone()[0]
        # Create first change log entry
        sql = 'INSERT INTO change_log_entries (parent_id, action, changed_by_id, changed_at) VALUES (%s, %s, %s, now()) RETURNING id'
        c.execute(sql, (change_log_id, 'Create', 1))

        # Create workflow log
        sql = 'INSERT INTO workflow_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        workflow_log_id = c.fetchone()[0]
        # Create first workflow log entry
        sql = 'INSERT INTO workflow_log_entries (parent_id, action, capacity, changed_by_id, changed_at) VALUES (%s, %s, %s, %s, now()) RETURNING id'
        c.execute(sql, (workflow_log_id, 'Create', 'Owner', 1))

        # Create vendor_part
        if not cbe:
            cbe = 0
        if not unc:
            unc = 0
        pbe = float(cbe) * (1 + (float(unc) / 100))
        state = 'Planned'
        owner_id = 1  # No need for owner_name or anything, it's always Sean with ID 1
        created_by_id = 1  # Same as above
        sql = 'INSERT INTO vendor_parts (part_number, name, current_best_estimate, uncertainty, predicted_best_estimate, material_id, material_specification_id, state, summary, owner_id, created_by_id, created_at, project_id, vendor_id, change_log_id, workflow_log_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), %s, %s, %s, %s) RETURNING id'
        c.execute(sql, (part_number, name, cbe, unc, pbe, material_id, material_spec_id, state, summary, owner_id, created_by_id, project_id, vendor_id, change_log_id, workflow_log_id))
        vendor_part_id = c.fetchone()[0]

        # Create change log
        sql = 'INSERT INTO change_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        change_log_id = c.fetchone()[0]
        # Create first change log entry
        sql = 'INSERT INTO change_log_entries (parent_id, action, changed_by_id, changed_at) VALUES (%s, %s, %s, now()) RETURNING id'
        c.execute(sql, (change_log_id, 'Create', 1))

        # Create build
        manufacturer_id = vendor_id
        sql = 'INSERT INTO vendor_builds (build_identifier, vendor_part_id, vendor_id, manufacturer_id, owner_id, created_by_id, created_at, change_log_id) VALUES (%s, %s, %s, %s, %s, %s, now(), %s) RETURNING id'
        c.execute(sql, (build_identifier, vendor_part_id, vendor_id, manufacturer_id, owner_id, created_by_id, change_log_id))
        vendor_build_id = c.fetchone()[0]

        # Create change log
        sql = 'INSERT INTO change_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        change_log_id = c.fetchone()[0]
        # Create first change log entry
        sql = 'INSERT INTO change_log_entries (parent_id, action, changed_by_id, changed_at) VALUES (%s, %s, %s, now()) RETURNING id'
        c.execute(sql, (change_log_id, 'Create', 1))

        # Create workflow log
        sql = 'INSERT INTO workflow_logs DEFAULT VALUES RETURNING id'
        c.execute(sql)
        workflow_log_id = c.fetchone()[0]
        # Create first workflow log entry
        sql = 'INSERT INTO workflow_log_entries (parent_id, action, capacity, changed_by_id, changed_at) VALUES (%s, %s, %s, %s, now()) RETURNING id'
        c.execute(sql, (workflow_log_id, 'Create', 'Owner', 1))

        # Create vendor_product
        state = 'Planned'
        if product_type == 'LOT':
            serial_number = 'L{0}'.format(serial_number)
        else:
            product_type = 'STOCK'
            serial_number = 'STCK'
        measured_mass = pbe  # NB! Major cheat!
        sql = 'INSERT INTO vendor_products (serial_number, vendor_part_id, summary, state, product_type, measured_mass, hardware_type_id, owner_id, created_by_id, created_at, project_id, vendor_build_id, change_log_id, workflow_log_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, now(), %s, %s, %s, %s) RETURNING id'
        c.execute(sql, (serial_number, vendor_part_id, summary, state, product_type, measured_mass, hardware_type_id, owner_id, created_by_id, product_project_id, vendor_build_id, change_log_id, workflow_log_id))


def insert_designs(c, wb):
    sheet = wb.get_sheet_by_name('DESIGNS')
    # ['Design Number', 'Name / Description', 'Summary', 'Owner', 'Project', 'CBE', 'UNC', 'Material', 'Matrl Spec']
    # ['id', 'design_number', 'revision', 'name', 'state', 'summary', 'notes', 'owner_id', 'created_by_id', 'created_at', 'project_id', 'export_control', 'thumbnail_id', 'revision_log_id', 'change_log_id']
    # ['id', 'part_identifier', 'current_best_estimate', 'uncertainty', 'predicted_best_estimate', 'design_id', 'material_id', 'material_specification_id', 'inseparable_component', 'name']
    for row in sheet.iter_rows(row_offset=1):
        design_number = row[0].value
        name = row[1].value
        summary = row[2].value
        project_name = row[4].value
        cbe = row[5].value
        unc = row[6].value
        material_name = row[7].value
        material_spec_name = row[8].value

        # Get project ID
        project_id = None
        if project_name:
            sql = 'SELECT id FROM projects WHERE name = %s'
            c.execute(sql, (project_name,))
            project_id = c.fetchone()[0]
        # Get material ID
        material_id = None
        if material_name:
            sql = 'SELECT id FROM materials WHERE name = %s'
            c.execute(sql, (material_name,))
            material_id = c.fetchone()[0]
        # Get material_spec ID
        material_spec_id = None
        if material_spec_name:
            sql = 'SELECT id FROM material_specifications WHERE name = %s'
            c.execute(sql, (material_spec_name,))
            material_spec_id = c.fetchone()[0]

        # Create design
        revision = 'A'
        state = 'Planned'
        owner_id = 1  # No need for owner_name or anything, it's always Sean with ID 1
        created_by_id = 1  # Same as above
        if design_number:
            # Create revision log
            sql = 'INSERT INTO revision_logs DEFAULT VALUES RETURNING id'
            c.execute(sql)
            revision_id = c.fetchone()[0]
            # Create first revision log entry
            sql = 'INSERT INTO revision_log_entries (parent_id, revision, reason, revisioned_by_id, revisioned_at) VALUES (%s, %s, %s, %s, now()) RETURNING id'
            c.execute(sql, (revision_id, 'A', 'Initial Release', 1))

            # Create change log
            sql = 'INSERT INTO change_logs DEFAULT VALUES RETURNING id'
            c.execute(sql)
            change_log_id = c.fetchone()[0]
            # Create first change log entry
            sql = 'INSERT INTO change_log_entries (parent_id, action, changed_by_id, changed_at) VALUES (%s, %s, %s, now()) RETURNING id'
            c.execute(sql, (change_log_id, 'Create', 1))

            # Create workflow log
            sql = 'INSERT INTO workflow_logs DEFAULT VALUES RETURNING id'
            c.execute(sql)
            workflow_log_id = c.fetchone()[0]
            # Create first workflow log entry
            sql = 'INSERT INTO workflow_log_entries (parent_id, action, capacity, changed_by_id, changed_at) VALUES (%s, %s, %s, %s, now()) RETURNING id'
            c.execute(sql, (workflow_log_id, 'Create', 'Owner', 1))

            # Create design
            sql = 'INSERT INTO designs (design_number, revision, name, state, summary, owner_id, created_by_id, created_at, project_id, change_log_id, revision_log_id, workflow_log_id) VALUES (%s, %s, %s, %s, %s, %s, %s, now(), %s, %s, %s, %s) RETURNING id'
            c.execute(sql, (design_number, revision, name, state, summary, owner_id, created_by_id, project_id, change_log_id, revision_id, workflow_log_id))
            design_id = c.fetchone()[0]

            # Create change log for part
            sql = 'INSERT INTO change_logs DEFAULT VALUES RETURNING id'
            c.execute(sql)
            change_log_id = c.fetchone()[0]
            # Create first change log entry
            sql = 'INSERT INTO change_log_entries (parent_id, action, changed_by_id, changed_at) VALUES (%s, %s, %s, now()) RETURNING id'
            c.execute(sql, (change_log_id, 'Create', 1))

            # Create part
            if not cbe:
                cbe = 0
            if not unc:
                unc = 0
            pbe = float(cbe) * (1 + (float(unc) / 100))
            part_identifier = '1'
            sql = 'INSERT INTO parts (part_identifier, name, current_best_estimate, uncertainty, predicted_best_estimate, design_id, material_id, material_specification_id, change_log_id, created_at, owner_id, created_by_id, inseparable_component) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, now(), %s, %s, %s) RETURNING id'
            c.execute(sql, (part_identifier, name, cbe, unc, pbe, design_id, material_id, material_spec_id, change_log_id, 1, 1, False))


def populate_data_sheets(wb):
    data_sheets = {}
    for sheet_name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheet_name)
        headers = []
        index = 1
        cell = sheet.cell(column=index, row=1)
        while cell.value:
            headers.append(cell.value.strip())
            index += 1
            cell = sheet.cell(column=index, row=1)
        data_sheets[sheet_name] = headers
    return data_sheets


def populate_data_tables(c, tables):
    data_tables = {}
    for table in tables:
        c.execute("select column_name from information_schema.columns where table_schema = 'public' and table_name='{0}'".format(table))
        data_tables[table] = [row[0] for row in c]
    return data_tables


def populate_expected_sheets():
    expected_criticality_headers = ['NAME', 'DESCRIPTION', 'ORDERING']
    expected_disposition_headers = ['NAME', 'DESCRIPTION', 'ORDERING']
    expected_hwtype_headers = ['NAME', 'DESCRIPTION', 'ORDERING']
    expected_project_headers = ['NAME', 'DESCRIPTION']
    expected_company_headers = ['NAME', 'Website', 'PRI Account #', 'Terms', 'PRI Alias / Username', 'Address', 'Notes']
    expected_material_headers = ['NAME', 'Description', 'Spec 1', 'Spec 2', 'Spec 3', 'Spec 4', 'Ordering']
    expected_specs_headers = ['Spec Number', 'Title', 'Owner']
    expected_vendor_parts_products_headers = ['Part Number', 'Name / Description', 'Summary', 'Vendor', 'Owner', 'Project', 'CBE', 'UNC', 'Material', 'Matrl Spec', 'Build', 'TYPE', 'S/L #', 'Hardware Type', 'Project', 'Owner']
    expected_design_headers = ['Design Number', 'Name / Description', 'Summary', 'Owner', 'Project', 'CBE', 'UNC', 'Material', 'Matrl Spec']
    expected_sheets = {}
    expected_sheets['CRITICALITY'] = expected_criticality_headers
    expected_sheets['DISPOSITION'] = expected_disposition_headers
    expected_sheets['HW TYPE'] = expected_hwtype_headers
    expected_sheets['PROJECT'] = expected_project_headers
    expected_sheets['COMPANY'] = expected_company_headers
    expected_sheets['MATERIAL'] = expected_material_headers
    expected_sheets['SPECS'] = expected_specs_headers
    expected_sheets['VENDOR Parts.Products'] = expected_vendor_parts_products_headers
    expected_sheets['DESIGNS'] = expected_design_headers
    # expected_sheets['Sheet1'] = []  # Sheet1 contains links etc, nothing to import
    return expected_sheets


def populate_expected_tablestructure():
    expected_criticalities_columns = ['id', 'name', 'description', 'ordering']
    expected_dispositions_columns = ['id', 'name', 'description', 'ordering']
    expected_hardware_types_columns = ['id', 'name', 'description', 'ordering']
    expected_projects_columns = ['id', 'name', 'description']
    expected_companies_columns = ['id', 'name', 'website', 'address', 'notes']
    expected_materials_columns = ['id', 'name', 'description']
    expected_material_specifications_columns = ['id', 'name', 'material_id']
    expected_specifications_columns = ['id', 'specification_number', 'name', 'owner_id']
    expected_vendor_parts_columns = ['id', 'part_number', 'name', 'state', 'summary', 'notes', 'owner_id', 'created_by_id', 'created_at', 'project_id', 'vendor_id', 'thumbnail_id', 'current_best_estimate', 'material_id', 'material_specification_id', 'predicted_best_estimate', 'uncertainty']
    expected_vendor_builds_columns = ['id', 'build_identifier', 'vendor_part_id', 'notes', 'purchase_order', 'vendor_id', 'manufacturer_id', 'owner_id', 'created_by_id', 'created_at', 'change_log_id']
    expected_vendor_products_columns = ['id', 'serial_number', 'vendor_part_id', 'summary', 'notes', 'state', 'product_type', 'measured_mass', 'hardware_type_id', 'owner_id', 'created_by_id', 'created_at', 'project_id', 'vendor_build_id', 'thumbnail_id', 'change_log_id']
    expected_designs_columns = ['id', 'design_number', 'revision', 'name', 'state', 'summary', 'notes', 'owner_id', 'created_by_id', 'created_at', 'project_id', 'export_control', 'thumbnail_id', 'revision_log_id', 'change_log_id']
    expected_parts_columns = ['id', 'part_identifier', 'current_best_estimate', 'uncertainty', 'predicted_best_estimate', 'design_id', 'material_id', 'material_specification_id', 'inseparable_component', 'name']
    expected_tables = {}
    expected_tables['criticalities'] = expected_criticalities_columns
    expected_tables['dispositions'] = expected_dispositions_columns
    expected_tables['hardware_types'] = expected_hardware_types_columns
    expected_tables['projects'] = expected_projects_columns
    expected_tables['companies'] = expected_companies_columns
    expected_tables['materials'] = expected_materials_columns
    expected_tables['material_specifications'] = expected_material_specifications_columns
    expected_tables['specifications'] = expected_specifications_columns
    expected_tables['vendor_parts'] = expected_vendor_parts_columns
    expected_tables['vendor_builds'] = expected_vendor_builds_columns
    expected_tables['vendor_products'] = expected_vendor_products_columns
    expected_tables['designs'] = expected_designs_columns
    expected_tables['parts'] = expected_parts_columns
    return expected_tables


if __name__ == '__main__':
   main()
